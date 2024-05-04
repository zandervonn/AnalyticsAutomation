import os
import time

import openpyxl

from Src.helpers.file_helpers import get_excel_csv_files
from Src.helpers.json_helpers import *


def split_json_list_columns(df, column, keys):
	"""
	Split the JSON or list stored in a column into separate columns for each specified key.

	Parameters:
	- df (pd.DataFrame): DataFrame containing the column with JSON or list.
	- column (str): Column name in DataFrame where the JSON or list is stored.
	- keys (list): List of keys to extract values for from the JSON or list.

	Returns:
	- pd.DataFrame: DataFrame with new columns for each key with values extracted.
	"""

	# Create new columns for each key
	for key in keys:
		df[f"{column}.{key}"] = None

	# Iterate over each row and process the data in the specified column
	for index, row in df.iterrows():
		# Check if the data is a dictionary or a list of dictionaries
		if isinstance(row[column], dict):
			items = [row[column]]  # Wrap the dictionary in a list
		elif isinstance(row[column], list):
			items = row[column]
		else:
			items = parse_json(row[column]) if pd.notna(row[column]) else []

		# Ensure items is a list of dicts
		if isinstance(items, list) and all(isinstance(item, dict) for item in items):
			# Extract values for each key
			for key in keys:
				# Join the values with line breaks and assign to the DataFrame
				df.at[index, f"{column}.{key}"] = "\n".join(str(item.get(key, '')) for item in items)

	# Optionally, drop the original column if no longer needed
	df.drop(columns=[column], inplace=True)

	return df

def csv_sheets_to_excel(csv_files, excel_file):
	"""
	Convert multiple CSV files into a single Excel workbook with each CSV file as a separate sheet.

	Parameters:
	- csv_files (list): List of paths to the CSV files.
	- excel_file (str): Path to save the Excel workbook.

	Returns:
	- None
	"""

	with pd.ExcelWriter(excel_file) as writer:
		for csv_file in csv_files:
			try:
				df = pd.read_csv(csv_file)
			except pd.errors.EmptyDataError:
				df = pd.DataFrame()  # Create an empty DataFrame if the CSV is empty
			# Use only the file name for the sheet name, removing invalid characters
			sheet_name = os.path.basename(csv_file).split('.')[0]
			sheet_name = sheet_name.replace('\\', '_').replace('/', '_').replace('*', '').replace('?', '').replace(':', '').replace('[', '').replace(']', '')
			# Shorten the sheet name to the Excel limit of 31 characters if necessary
			sheet_name = sheet_name[:31]
			df.to_excel(writer, sheet_name=sheet_name, index=False)
	print(f"CSV saved to {excel_file}")

def save_df_to_csv(df, file_name, safe_save=False):
	"""
	Save a DataFrame to a CSV file with optional retry logic if the file is locked.

	Parameters:
	- df (pd.DataFrame): DataFrame to save to CSV.
	- file_name (str): Path for the output CSV file.
	- safe_save (bool): If True, will retry saving if the file is locked.

	Returns:
	- None
	"""

	start_time = time.time()
	retry_interval=5
	timeout=120

	while True:
		try:
			# Try to save the DataFrame to a CSV file
			df.to_csv(file_name, encoding='utf-8-sig', index=False)
			print(f"CSV saved to {file_name}")
			break  # Exit the loop if the file is successfully saved
		except PermissionError as e:
			# Handle the case where the file is open and locked
			if safe_save:
				elapsed_time = time.time() - start_time
				if elapsed_time >= timeout:
					print(f"Failed to save CSV after {timeout} seconds. File may be open.")
					break  # Exit the loop after timeout
				print(f"File is locked, retrying in {retry_interval} seconds...")
				time.sleep(retry_interval)  # Wait before retrying
			else:
				print(f"Failed to save CSV: {e}")
				break
		except Exception as e:
			# Handle other exceptions
			print(f"An error occurred: {e}")
			break

def save_df_to_excel(df_or_dict, filename):
	"""
	Save a DataFrame or a dictionary of DataFrames to an Excel file.

	Parameters:
	- df_or_dict (pd.DataFrame | dict): Single DataFrame or dictionary of DataFrames to save.
	- filename (str): Path for the output Excel file.

	Returns:
	- None
	"""
	with pd.ExcelWriter(filename, engine='openpyxl') as writer:
		if isinstance(df_or_dict, pd.DataFrame):
			# If a single DataFrame is provided, save it to the first sheet
			if not df_or_dict.empty:
				df_or_dict.to_excel(writer, sheet_name='Sheet1', index=False)
			else:
				print("The input DataFrame is empty. No Excel file created.")
		elif isinstance(df_or_dict, dict):
			# If a dictionary of DataFrames is provided, save each DataFrame to a separate sheet
			for sheet_name, df in df_or_dict.items():
				if not isinstance(df, pd.DataFrame):
					raise TypeError(f"The value for key '{sheet_name}' is not a DataFrame. It is of type {type(df)}.")
				if not df.empty:
					df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Sheet name limited to 31 characters
				else:
					print(f"Skipping empty DataFrame for key '{sheet_name}'.")
		else:
			raise TypeError("Input must be a DataFrame or a dictionary of DataFrames.")

	print(f"Excel saved to {filename}")


def load_csv(path):
	"""
	Load a CSV file into a DataFrame with error handling.

	Parameters:
	- path (str): Path to the CSV file to load.

	Returns:
	- pd.DataFrame: Loaded data as DataFrame or empty DataFrame if loading fails.
	"""
	try:
		return pd.read_csv(path)
	except Exception as e:
		print(f"Warning: Failed to load {path}. Error: {e}")
		return pd.DataFrame()

def write_data_to_csv(data, file_path):
	"""
	Write a dictionary of data to a CSV file.

	Parameters:
	- data (dict): Data to write where keys are column headers and values are the row data.
	- file_path (str): Path for the output CSV file.

	Returns:
	- None
	"""
	with open(file_path, mode='w', newline='') as file:
		writer = csv.writer(file)
		writer.writerow(data.keys())
		writer.writerow(data.values())

def update_columns(df, raw_data):
	"""
	Update columns of a DataFrame with data from a dictionary of raw data.

	Parameters:
	- df (pd.DataFrame): DataFrame with columns to be updated.
	- raw_data (dict): Dictionary where keys are file names and values are DataFrames with update data.

	Returns:
	- pd.DataFrame: Updated DataFrame.
	"""
	for col in df.columns:
		period_count = col.count('.')
		if period_count >= 2:
			parts = col.split('.', 2)
			file, sheet, name = parts if period_count == 2 else parts[:2] + ['.'.join(parts[2:])]
			file += '.xlsx'  # Add the .xlsx extension to the file name
			if file in raw_data:
				if sheet in raw_data[file]:
					if name in raw_data[file][sheet].columns:
						print(f"Found column '{name}' in sheet '{sheet}' of file '{file}'. Updating...")
						df[col] = raw_data[file][sheet][name]
	return df

def update_files(raw_folder, update_folder):
	"""
	Update Excel files in a folder using data from another folder's Excel files.

	Parameters:
	- raw_folder (str): Path to the folder containing the raw data Excel files.
	- update_folder (str): Path to the folder containing Excel files to be updated.

	Returns:
	- None
	"""
	print("Updating files...")

	raw_files = get_excel_csv_files(raw_folder)
	update_files = get_excel_csv_files(update_folder)

	print("Number of output files found:", len(raw_files))
	print("Number of custom files found:", len(update_files))

	raw_data = load_excel_files_into_dict(raw_files)

	for update_file in update_files:
		# Load the workbook and iterate through each sheet
		wb = openpyxl.load_workbook(update_file, data_only=False)
		for sheet_name in wb.sheetnames:
			sheet = wb[sheet_name]
			df = pd.read_excel(update_file, sheet_name=sheet_name)
			updated_df = update_columns(df, raw_data)

			# Update the cells with values from the updated DataFrame
			for row in range(len(updated_df)):
				for col, value in enumerate(updated_df.iloc[row]):
					cell = sheet.cell(row=row + 2, column=col + 1)  # Adjust indexes for Excel (1-based indexing)
					# Only update the cell if it does not contain a formula
					if cell.value is None or not isinstance(cell.value, str) or not cell.value.startswith('='):
						cell.value = value

		# Save the workbook
		wb.save(update_file)

	print("Files updated successfully")
