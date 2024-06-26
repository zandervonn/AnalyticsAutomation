import openpyxl
import pandas as pd
import re
import os

import pytz
from bs4 import BeautifulSoup
from datetime import datetime
from dateutil.parser import parse, ParserError
from openpyxl.styles import NamedStyle
from Src.access import output_folder_path, final_output_path


def path_gen(*args):
	if len(args) == 4:
		branch, platform, data_type, file_format = args

		# Create the subfolder path based on the platform and data type
		subfolder_path = os.path.join(output_folder_path(), branch, platform, data_type)

		file_name = f"{platform}_{data_type}.{file_format}"

	elif len(args) == 3:
		platform, data_type, file_format = args

		# Create the subfolder path based on the platform and data type
		subfolder_path = os.path.join(output_folder_path(), platform, data_type)

		file_name = f"{platform}_{data_type}.{file_format}"

	elif len(args) == 2:
		branch, platform = args

		# Create the subfolder path based on the platform
		subfolder_path = os.path.join(output_folder_path(), branch, 'output', platform)

		# Construct the file name
		file_name = f"{platform}_output.xlsx"

	else:
		raise ValueError("Invalid number of arguments. Expected 1, 3, or 4.")

	# Ensure the subfolder exists
	os.makedirs(subfolder_path, exist_ok=True)

	# Construct the full file path
	file_path = os.path.join(subfolder_path, file_name)

	return file_path

def find_path_upwards(target_relative_path, start_path=None):
	"""
    Searches for a target relative path upwards from the start path until the root is reached.

    :param target_relative_path: The relative path to search for.
    :param start_path: The starting directory for the search. If None, uses the directory of the current file.
    :return: The absolute path to the target if found, None otherwise.
    """
	if start_path is None:
		start_path = os.path.dirname(__file__)

	current_path = os.path.abspath(start_path)
	root_path = os.path.abspath(os.sep)

	while current_path != root_path:
		potential_target_path = os.path.join(current_path, target_relative_path)
		if os.path.exists(potential_target_path):
			return potential_target_path
		current_path = os.path.dirname(current_path)

	return None

def get_header_list(list_name):
	excel_file_path = find_path_upwards(r'config/headers.xlsx')
	wb = openpyxl.load_workbook(excel_file_path)

	# Get the formatting sheet and the ignore format
	formatting_sheet = wb['formatting']
	ignore_format = formatting_sheet['A1'].fill

	# Get the headers sheet
	headers_sheet = wb['headers']

	for row in headers_sheet.iter_rows():
		if row[0].value == list_name:
			return [cell.value for cell in row[1:] if cell.value and not cell.fill.start_color.index == ignore_format.start_color.rgb]
	return []

def load_mapping(file_path):
	mapping = {}
	with open(file_path, 'r') as file:
		for line in file:
			parts = line.strip().split('=')
			if len(parts) == 2:
				key, value = parts[0].strip(), parts[1].strip()
				mapping[key] = value
	return mapping

def remove_html_tags(text):
	# Remove HTML tags
	text = BeautifulSoup(text, "html.parser").get_text()
	# Remove non-printable characters
	text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
	return text

def read_and_clean_data(file_path):
	try:
		file_extension = os.path.splitext(file_path)[1].lower()
		if file_extension == '.csv':
			df = pd.read_csv(file_path)
			df = df.applymap(lambda x: remove_html_tags(x) if isinstance(x, str) else x)
			return {'Sheet1': df}
		elif file_extension in ['.xls', '.xlsx']:
			xls = pd.ExcelFile(file_path)
			dfs = {sheet_name: xls.parse(sheet_name).applymap(lambda x: remove_html_tags(x) if isinstance(x, str) else x) for sheet_name in xls.sheet_names}
			return dfs
		else:
			print(f"Unsupported file type: {file_path}")
			return None
	except Exception as e:
		print(f"Error processing {file_path}: {e}")
		return None

def save_data_to_excel(df, writer, sheet_name):
	# Write the DataFrame to the ExcelWriter
	df.to_excel(writer, sheet_name=sheet_name, index=False)

	# Get the workbook and worksheet for formatting
	workbook = writer.book
	worksheet = writer.sheets[sheet_name]

	# Define a date style format using openpyxl
	date_style = NamedStyle(name='custom_date_style', number_format='YYYY-MM-DD')

	# Apply the style to each cell in the date columns
	date_columns = [col for col in df.columns if df[col].dtype in ['datetime64[ns]', 'datetime64[ns, tz]']]
	for col in date_columns:
		column_letter = openpyxl.utils.get_column_letter(df.columns.get_loc(col) + 1)  # get_loc is zero indexed, openpyxl is 1 indexed
		for row in range(2, len(df) + 2):  # Start at 2 to skip the header row
			worksheet[f'{column_letter}{row}'].style = date_style

def create_date_saved_df(all_data_columns):
	date_saved = 'Date Saved: ' + datetime.now().strftime('%Y-%m-%d')
	return pd.DataFrame([[date_saved] + [''] * (all_data_columns - 1)])

def files_to_excel(file_paths, output_excel_path):
	with pd.ExcelWriter(output_excel_path, engine='openpyxl', datetime_format='YYYY-MM-DD') as writer:
		for file_path in file_paths:
			dfs = read_and_clean_data(file_path)
			if dfs:
				if len(dfs) == 1:
					# Single sheet or CSV case, use the base file name as the sheet name
					df = list(dfs.values())[0]
					sheet_name = os.path.basename(file_path).split('.')[0][:31]
					save_data_to_excel(df, writer, sheet_name)
				else:
					# Multiple sheets case, append sheet names
					for sheet_name, df in dfs.items():
						safe_sheet_name = (os.path.basename(file_path).split('.')[0] + '_' + sheet_name)[:31]
						save_data_to_excel(df, writer, safe_sheet_name)

	print(f"Files saved to {output_excel_path}")

def update_worksheet_data(ws, col, df, column, update_header, file, sheet):
	"""
	Updates data or headers in a worksheet based on the specified column information,
	including date formatting if applicable.

	Parameters:
	- ws: The worksheet to update.
	- col: The column index to start from.
	- df: DataFrame containing the data.
	- column: Specific column name or 'all' to update.
	- update_header: Boolean to determine if headers should be updated.
	- file: File name used in messaging.
	- sheet: Sheet name used in messaging.
	"""
	# Apply date formatting based on column headers in the worksheet
	df = format_date_columns(df)

	if column == "all":
		replace_entire_sheet(ws, df, col, update_header)
	elif column in df.columns:
		replace_column_data(ws, col, df[column], update_header)
	else:
		print(f"Column '{column}' not found in data for key: '{file}.{sheet}'")
		print(f"Available columns in '{sheet}' of '{file}': {list(df.columns)}")

def format_date_columns(df):
	"""
	Format columns in the DataFrame by checking if the second cell in each column is a date.
	Only formats columns where the second cell can be parsed as a date, and ensures that the cell exists.
	Ensures that formatted dates do not include a time component.

	Parameters:
	- df (pd.DataFrame): DataFrame to format.

	Returns:
	- pd.DataFrame: DataFrame with date columns formatted where applicable.
	"""
	for col in df.columns:
		# Check if there is at least one data row beyond the header
		if len(df[col]) > 1 and pd.notna(df[col].iloc[1]):  # Use iloc to safely access the second row
			try:
				# Attempt to parse the second row (first row after the header)
				parsed_date = parse(df[col].iloc[1], fuzzy=False)
				# If parse is successful, format the whole column without time component
				df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d').fillna(df[col])
			except (ParserError, ValueError, TypeError):
				# If parsing fails, do not treat as a date column
				continue
	return df

def is_convertible_to_date(value):
	"""
	Helper function to check if a value can be parsed as a date.

	Parameters:
	- value: The value to check.

	Returns:
	- bool: True if the value can be parsed as a date, False otherwise.
	"""
	try:
		parse(value, fuzzy=False)
		return True
	except (ParserError, ValueError):
		return False

# def is_date_format(number_format):
# 	# Define a list of date formats you expect to consider as date columns
# 	date_formats = ["yy-mm-dd", "yyyy-mm-dd", "dd-mm-yy", "dd-mm-yyyy", "m/d/yy", "m/d/yyyy", "d-mmm-yy", "yyyy-mm-dd h:mm:ss"]
# 	# Print what is being matched against for troubleshooting
# 	print(f"Checking formats in: {date_formats}")
# 	return any(fmt in number_format for fmt in date_formats)

def safe_assign_to_excel(ws, row, col, value):
	"""
	Safely assign a value to an Excel cell, ensuring that the value is compatible with Excel.
	"""
	try:
		ws.cell(row=row, column=col).value = value
	except ValueError:
		ws.cell(row=row, column=col).value = str(value)  # Convert to string if not directly assignable

def replace_column_data(ws, col, data, update_header):
	data_start_row = 2 if update_header else 3
	for i, value in enumerate(data, start=data_start_row):
		safe_assign_to_excel(ws, i, col, value)

	if update_header:
		ws.cell(row=1, column=col).value = data.name.capitalize()

def todays_output_folder(timezone='Pacific/Auckland'):
	# Create a subfolder in the output directory for today's date
	# Set the timezone
	tz = pytz.timezone(timezone)
	today = datetime.now(tz).strftime("%Y-%m-%d")
	date_output_folder = os.path.join(final_output_path(), today)
	os.makedirs(date_output_folder, exist_ok=True)
	return date_output_folder

def copy_template_num_formatting(template_ws, output_ws):
	for col in range(1, template_ws.max_column + 1):
		# Get the numeric format from the template column's first data cell
		template_format = template_ws.cell(row=2, column=col).number_format

		# Check if the format is a date type
		if 'd' in template_format.lower() and 'm' in template_format.lower() and 'y' in template_format.lower():
			template_format = 'yyyy-mm-dd'

		# Apply the numeric format to the entire output column
		for row in range(2, output_ws.max_row + 1):
			output_ws.cell(row=row, column=col).number_format = template_format

def update_template_files(template_folder, data_folder, output_folder):
	# Load data from the specified folder and print available keys
	data_files = get_excel_csv_files(data_folder)
	data_dict = load_files_into_dict(data_files)

	# Load template files
	template_files = get_excel_csv_files(template_folder)

	# Process each template file
	for template_file in template_files:
		process_template_file(template_file, data_dict, output_folder)

def process_template_file(template_file, data_dict, output_folder):
	wb = openpyxl.load_workbook(template_file)
	for sheet_name in wb.sheetnames:
		process_sheet(wb[sheet_name], data_dict, wb[sheet_name])

	# Save the workbook to the specified output folder with a dated filename
	save_workbook(wb, template_file, output_folder)

def process_sheet(template_ws, data_dict, output_ws):
	header = template_ws[2]  # Assuming second row contains references
	for col, cell in enumerate(header, start=1):
		update_cell(output_ws, col, cell, data_dict)
	output_ws.delete_rows(2)  # Remove the reference row after updates

	# Copy numeric formatting from template to output
	copy_template_num_formatting(template_ws, output_ws)

def update_cell(output_ws, col, cell, data_dict):
	ref = str(cell.value).strip().lower() if cell.value else ""
	if not ref or '.' not in ref:
		return  # Skip cells without proper references

	update_header = ref.endswith('_headers')
	ref = ref[:-8] if update_header else ref  # Remove '_headers' suffix if needed
	parts = ref.split('.')
	if len(parts) < 3:
		print(f"Reference in cell {cell.coordinate} is incomplete: '{cell.value}'")
		return

	file, sheet, column = (part.strip() for part in parts)
	key = f"{file}.{sheet}".lower()
	if key not in data_dict:
		print(f"Could not find data for key: '{key}'")
		return

	df = data_dict[key]
	update_worksheet_data(output_ws, col, df, column, update_header, file, sheet)

def save_workbook(wb, template_file, output_folder):
	# Set the focused cell to the top-left cell on all sheets
	for sheet in wb.worksheets:
		sheet.sheet_view.selection[0].activeCell = 'A1'
		sheet.sheet_view.selection[0].sqref = 'A1'

	# Set the focused sheet to the first sheet
	wb.active = 0

	today = datetime.now().strftime("%Y-%m-%d")
	file_name, file_extension = os.path.splitext(os.path.basename(template_file))
	output_path = os.path.join(output_folder, f"{file_name}_{today}{file_extension}")
	wb.save(output_path)
	print(f"Saved processed file to: {output_path}")

def capitalize_headers(headers):
	""" Capitalize the first letter of each header """
	return [header.capitalize() for header in headers]

def replace_entire_sheet(ws, df, start_col, update_header):
	""" Replace the entire sheet's data and optionally update headers """
	data_start_row = 2 if update_header else 3

	# Clear existing data from the sheet
	for row in ws.iter_rows(min_row=data_start_row, min_col=start_col, max_col=start_col+len(df.columns)-1):
		for cell in row:
			cell.value = None

	# Insert new data
	for idx, row in enumerate(df.itertuples(index=False), start=data_start_row):
		for j, value in enumerate(row, start=start_col):
			ws.cell(row=idx, column=j).value = value

	# Update headers if required
	if update_header:
		headers = capitalize_headers(list(df.columns))
		for i, header in enumerate(headers, start=start_col):
			ws.cell(row=1, column=i).value = header

def indicate_missing(ws, row, col, file, sheet, column=None):
	if column:
		message = f"MISSING {file}.{sheet}.{column}"
	else:
		message = f"MISSING {file}.{sheet}"
	ws.cell(row=row, column=col, value=message)

def load_files_into_dict(files):
	"""
	Load multiple data files (Excel and CSV) into a dictionary of DataFrames.
	Keys are formatted as 'filename.sheetname' for Excel files and 'filename.filename' for CSV files.
	"""
	data_dict = {}
	for file in files:
		file_name, file_ext = os.path.splitext(file)
		base_file_key = os.path.basename(file_name).lower().strip()

		try:
			if file_ext in ['.xlsx', '.xls']:
				wb = openpyxl.load_workbook(file, data_only=True)
				for sheet in wb.sheetnames:
					df = pd.read_excel(file, sheet_name=sheet)
					df.columns = df.columns.str.strip().str.lower()
					key = f"{base_file_key}.{sheet.lower().strip()}"
					data_dict[key] = df
			elif file_ext == '.csv':
				df = pd.read_csv(file)
				df.columns = df.columns.str.strip().str.lower()
				# For CSV files, assume sheet name is same as file name
				key = f"{base_file_key}.{base_file_key}"
				data_dict[key] = df
		except Exception as e:
			print(f"Error loading file {file}: {e}")
	return data_dict

def get_excel_files_from_folder(folder):
	"""
	Retrieve a list of all Excel files in a folder.

	Parameters:
	- folder (str): Path to the folder to search for Excel files.

	Returns:
	- list: List of paths to the Excel files.
	"""
	excel_files = []
	for root, _, files in os.walk(folder):
		for file in files:
			if file.endswith('.xlsx'):
				excel_files.append(os.path.join(root, file))
	return excel_files

def get_excel_csv_files(folder):
	"""
	Find and print paths of all Excel and CSV files in a folder.

	Parameters:
	- folder (str): Path to the folder to search for Excel and CSV files.

	Returns:
	- list: List of paths to the Excel and CSV files.
	"""
	excel_files = []
	for root, _, files in os.walk(folder):
		for file in files:
			if file.endswith(('.xlsx', '.xls', '.csv')):
				file_path = os.path.join(root, file)
				excel_files.append(file_path)
	return excel_files


def get_all_files(folder):
	"""
	Create a dictionary with file names as keys and full paths as values for all Excel and CSV files in a directory.
	"""
	files_dict = {}
	for root, _, files in os.walk(folder):
		for file in files:
			if file.endswith(('.xlsx', '.xls', '.csv')):
				file_key = file.lower().split('.')[0]
				files_dict[file_key] = os.path.join(root, file)
	return files_dict

def load_data(file_path, sheet_name=None):
	"""
	Load data from an Excel or CSV file into a DataFrame.
	"""
	ext = os.path.splitext(file_path)[1]
	if ext in ['.xlsx', '.xls']:
		return pd.read_excel(file_path, sheet_name=sheet_name)
	elif ext == '.csv':
		return pd.read_csv(file_path)
	else:
		raise ValueError(f"Unsupported file extension: {ext}")


def get_latest_marketing_file(folder_path, branch, input_date):
	"""
	Retrieve the marketing file for the given branch with the latest date preceding or equal to the input date.

	Parameters:
	- folder_path (str): The path to the folder containing the files.
	- branch (str): The branch name to include in the filename.
	- input_date (str): The input date in 'YYYY-MM-DD' format.

	Returns:
	- pd.DataFrame: The DataFrame of the extracted marketing file containing the sheet with 'current' in its name.
	"""
	# Parse input_date and strip time component if present
	input_date = parse(input_date).date()
	files = get_excel_csv_files(folder_path)

	marketing_files = []
	for file in files:
		match = re.search(r'marketing department ' + re.escape(branch) + r'_(\d{4}-\d{2}-\d{2})', file, re.IGNORECASE)
		if match:
			file_date = datetime.strptime(match.group(1), '%Y-%m-%d').date()
			if file_date <= input_date:
				marketing_files.append((file_date, file))

	if not marketing_files:
		raise FileNotFoundError("No marketing files found preceding or on the input date.")

	latest_file = max(marketing_files, key=lambda x: x[0])[1]
	print(f"Latest file selected: {latest_file}")

	# Load the file and extract the sheet containing 'current' in its name
	with pd.ExcelFile(latest_file) as xls:
		for sheet_name in xls.sheet_names:
			if 'current' in sheet_name.lower():
				return pd.read_excel(xls, sheet_name=sheet_name)
		raise ValueError("No sheet containing 'current' found in the file.")